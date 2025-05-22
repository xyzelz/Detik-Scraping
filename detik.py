import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
import re
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

MAX_URLS = 6000
IN_DEPTH_KEYWORDS = ['analisis', 'mengapa', 'bagaimana','opini', 'laporan khusus', 'investigasi', 'mendalam', 'penjelasan','penyebab']
WORD_COUNT_THRESHOLD = 300

def collect_search_urls(keyword_url):
    session = requests.Session()
    urls = []
    page = 1

    while len(urls) < MAX_URLS:
        try:
            url = f'https://www.detik.com/search/searchall?query={keyword_url}&sortby=time&page={page}'
            response = session.get(url, timeout=15)

            if "Tak ada hasil pencarian" in response.text:
                break

            soup = BeautifulSoup(response.text, 'html.parser')
            articles = soup.find_all('article')

            if not articles:
                break

            for article in articles:
                title = article.find('h3', class_='media__title')
                if not title:
                    continue

                link_tag = title.find('a')
                if not link_tag or not link_tag.has_attr('href'):
                    continue

                link = link_tag['href']
                media = article.find('h2', class_='media__subtitle')
                media = media.get_text(strip=True) if media else 'detikcom'

                urls.append((link, media))

            print(f"📄 Halaman {page} selesai, URL terkumpul: {len(urls)}", end='\r')
            page += 1
            time.sleep(0.3)

        except Exception as e:
            print(f"⚠️ Error koleksi URL, halaman {page}: {e}")
            time.sleep(1)
            continue

    return urls

def is_data_complete(data):
    required_fields = ['Judul', 'Isi Berita', 'Tanggal Publikasi', 'Media', 'URL', 'Jenis Berita']
    for field in required_fields:
        if field not in data or not data[field] or (isinstance(data[field], str) and data[field].strip() == ''):
            return False
    if len(data['Isi Berita'].split()) < 20:
        return False
    return True

def process_article(url_info):
    url, media = url_info
    try:
        with requests.Session() as session:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0 Safari/537.36'
            }
            response = session.get(url, timeout=20, headers=headers)
            if response.status_code != 200:
                return None

            soup = BeautifulSoup(response.text, 'html.parser')

            title = soup.find('h1', class_='detail__title') or soup.find('h1')
            if not title:
                return None
            title_text = title.get_text(strip=True)
            if not title_text:
                return None

            date_element = soup.find('div', class_='detail__date') or soup.find(class_=lambda x: x and ('date' in x.lower() or 'time' in x.lower()))
            if not date_element:
                return None
            date_text = date_element.get_text(strip=True)
            if not date_text:
                return None

            main_content = soup.find('div', class_='detail__body') or soup.find('div', class_=lambda x: x and ('body' in x.lower() or 'content' in x.lower())) or soup.find('article')
            if not main_content:
                return None
            content_elements = main_content.find_all(['h2', 'h3', 'p'])
            if not content_elements:
                return None
            content = '\n'.join([p.get_text(strip=True) for p in content_elements])
            if not content:
                return None
            word_count = len(content.split())
            if word_count < 20:
                return None

            is_in_depth = (
                word_count > WORD_COUNT_THRESHOLD or
                any(kw in content.lower() for kw in IN_DEPTH_KEYWORDS) or
                any(kw in title_text.lower() for kw in IN_DEPTH_KEYWORDS) or
                'detikNews' not in media
            )

            article_data = {
                'Judul': title_text,
                'Isi Berita': content,
                'Tanggal Publikasi': date_text,
                'Media': media,
                'URL': url,
                'Jenis Berita': "In-depth Reporting" if is_in_depth else "Straight News",
                'Jumlah Kata': word_count
            }

            if is_data_complete(article_data):
                return article_data
            else:
                return None

    except Exception as e:
        print(f"⚠️ Error proses {url}: {e}")
        return None

def autofit_columns(ws, max_width=80):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        adjusted_width = min(max_len + 3, max_width)
        ws.column_dimensions[col_letter].width = adjusted_width


def style_worksheet(ws, sheet_title):
    max_col = ws.max_column
    max_row = ws.max_row
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"📄 {sheet_title} Report"
    title_cell.font = Font(size=30, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    header_fill = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in ws.iter_rows(min_row=3, max_row=max_row):
        for cell in row:
            if cell.row % 2 == 0:
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    autofit_columns(ws)

if __name__ == "__main__":
    print("📰 DETIK.COM SCRAPER NEWS\n")

    keyword = input("🔍 Masukkan topik pencarian: ").strip()
    keyword_url = keyword.replace(" ", "+")

    try:
        target_straight = int(input("📊 Masukkan jumlah artikel Straight News yang diinginkan: ").strip())
    except ValueError:
        print("⚠️ Input tidak valid, menggunakan default 500")
        target_straight = 500

    try:
        target_indepth = int(input("📘 Masukkan jumlah artikel In-depth Reporting yang diinginkan: ").strip())
    except ValueError:
        print("⚠️ Input tidak valid, menggunakan default 500")
        target_indepth = 500

    print(f"\n🔎 Mengumpulkan URL untuk topik: '{keyword}'...\n")
    article_urls = collect_search_urls(keyword_url)
    print(f"✅ Total URL terkumpul: {len(article_urls)}")

    straight_news = []
    in_depth_reporting = []
    processed_count = 0

    print("⚙️ Memulai pemrosesan artikel...\n")

    with ThreadPoolExecutor(max_workers=25) as executor:
        futures = [executor.submit(process_article, url) for url in article_urls]

        for future in as_completed(futures):
            processed_count += 1
            result = future.result()

            if result:
                if result['Jenis Berita'] == 'Straight News' and len(straight_news) < target_straight:
                    straight_news.append(result)
                elif result['Jenis Berita'] == 'In-depth Reporting' and len(in_depth_reporting) < target_indepth:
                    in_depth_reporting.append(result)

            print(f"🔄 {processed_count}/{len(article_urls)} | 📰 Straight: {len(straight_news)}/{target_straight} | 📘 In-depth: {len(in_depth_reporting)}/{target_indepth}", end='\r')

            if len(straight_news) >= target_straight and len(in_depth_reporting) >= target_indepth:
                print("\n🎯 Target artikel berhasil terpenuhi!")
                break

    straight_news = straight_news[:target_straight]
    in_depth_reporting = in_depth_reporting[:target_indepth]

    print(f"\n📊 Hasil akhir:")
    print(f"   📰 Straight News: {len(straight_news)}/{target_straight}")
    print(f"   📘 In-depth Reporting: {len(in_depth_reporting)}/{target_indepth}")

    if len(straight_news) < target_straight or len(in_depth_reporting) < target_indepth:
        print("\n⚠️ PERINGATAN: Target jumlah artikel tidak tercapai!")
        if len(straight_news) < target_straight:
            print(f"   ➤ Kurang {target_straight - len(straight_news)} artikel Straight News 📰")
        if len(in_depth_reporting) < target_indepth:
            print(f"   ➤ Kurang {target_indepth - len(in_depth_reporting)} artikel In-depth 📘")
        print("💡 Saran: Coba gunakan kata kunci yang lebih umum atau kurangi batas jumlah artikel.\n")

    if len(straight_news) > 0 or len(in_depth_reporting) > 0:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f'detik_{keyword_url}_{timestamp}.xlsx'

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            if straight_news:
                df = pd.DataFrame(straight_news)
                df.to_excel(writer, sheet_name='Straight News', index=False, startrow=1)
            if in_depth_reporting:
                df = pd.DataFrame(in_depth_reporting)
                df.to_excel(writer, sheet_name='In-depth Reporting', index=False, startrow=1)

        wb = load_workbook(filename)
        if 'Straight News' in wb.sheetnames:
            style_worksheet(wb['Straight News'], "Straight News")
        if 'In-depth Reporting' in wb.sheetnames:
            style_worksheet(wb['In-depth Reporting'], "In-depth Reporting")
        wb.save(filename)

print(f"\n💾 Data berhasil disimpan dalam file: {filename}")
print(f"📂 Sheet 1: Straight News ({len(straight_news)} artikel)")
print(f"📂 Sheet 2: In-depth Reporting ({len(in_depth_reporting)} artikel)")
